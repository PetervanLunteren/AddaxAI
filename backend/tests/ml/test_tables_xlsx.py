"""Tests for the tables_xlsx postprocess output module.

The row schemas are covered by the export-layer tests; this wrapper just
makes sure one workbook with the Files and Detections sheets lands at the
right path, trimmed to the folder-run column set.
"""

from pathlib import Path

import pytest

from app.ml.postprocessing_outputs.tables_xlsx import (
    XLSX_FILENAME,
    write_tables_xlsx,
)
from tests.conftest import (
    make_deployment,
    make_detection,
    make_file,
    make_project,
)

# XLSX files are ZIP archives. ZIP magic is "PK\x03\x04".
_XLSX_MAGIC = b"PK\x03\x04"


def _write_placeholder(path: Path) -> str:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(b"x")
    return str(path)



def test_writes_two_sheet_workbook(db, tmp_path):
    from openpyxl import load_workbook

    project = make_project(db, name="xlsx-basic")
    dep = make_deployment(db, project_id=project.id)
    file = make_file(
        db,
        deployment_id=dep.id,
        file_path=_write_placeholder(tmp_path / "src" / "IMG.jpg"),
        observation_type="animal",
    )
    make_detection(
        db,
        file_id=file.id,
        category="animal",
        confidence=0.9,
        label="dog",
        bbox_x=0.1,
        bbox_y=0.1,
        bbox_width=0.2,
        bbox_height=0.2,
    )

    target = tmp_path / "out"
    result = write_tables_xlsx(db, project.id, target)

    assert result.output_path.endswith(XLSX_FILENAME)
    output_path = target / XLSX_FILENAME
    assert output_path.is_file()
    with open(output_path, "rb") as f:
        assert f.read(4) == _XLSX_MAGIC

    wb = load_workbook(output_path)
    assert wb.sheetnames == ["Files", "Detections"]

    # Same trimmed column set as the folder-run CSVs.
    for sheet in ("Files", "Detections"):
        headers = [c.value for c in wb[sheet][1]]
        assert "deployment_id" not in headers
        assert "notes" not in headers
        assert "file_id" in headers
        assert "event_id" in headers


def test_files_sheet_keeps_the_species_columns(db, tmp_path):
    """The XLSX writer is wired independently of the CSV one, so the folder-run
    Files sheet needs its own pin that the one label per file is present."""
    from openpyxl import load_workbook

    project = make_project(db, name="xlsx-species")
    dep = make_deployment(db, project_id=project.id)
    file = make_file(
        db,
        deployment_id=dep.id,
        file_path=_write_placeholder(tmp_path / "src" / "IMG.jpg"),
        observation_type="animal",
    )
    make_detection(
        db,
        file_id=file.id,
        category="animal",
        confidence=0.9,
        label="red fox",
        scientific_name="Vulpes vulpes",
        common_name="Red fox",
    )

    target = tmp_path / "out"
    write_tables_xlsx(db, project.id, target)

    wb = load_workbook(target / XLSX_FILENAME)
    rows = list(wb["Files"].values)
    headers, row = list(rows[0]), list(rows[1])
    species = [
        row[headers.index(name)]
        for name in ("classification_label", "scientific_name", "common_name")
    ]
    assert species == ["red fox", "Vulpes vulpes", "Red fox"]
    assert "taxon_family" in headers


def test_unknown_project_raises(db, tmp_path):
    with pytest.raises(ValueError, match="not found"):
        write_tables_xlsx(db, "no-such-id", tmp_path / "out")
