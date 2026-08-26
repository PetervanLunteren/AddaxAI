"""Tests for the /api/projects/{project_id}/export endpoints."""

from __future__ import annotations

import asyncio
import csv
import io
import json
import sqlite3
import tempfile
import uuid
import zipfile
from datetime import date, datetime
from unittest.mock import patch

import pytest

from app.api.crud import export_formats
from app.models import File
from app.models.label_taxonomy import LabelTaxonomy
from tests.conftest import (
    make_deployment,
    make_detection,
    make_event_with_files,
    make_file,
    make_project,
    make_site,
)


def _run_camtrap_dp_export(client, db, project_id: str):
    """Helper: drive the full prepare → worker → download cycle.

    The prepare endpoint registers the worker with ws_manager and waits
    for a frontend "ready" signal before running it. In tests we have no
    WebSocket, so we kick the worker directly (after patching its
    `get_db` to use a session bound to the test engine) and then GET
    the download endpoint.

    The worker's `finally: db.close()` would otherwise detach the
    fixture's `db` instance and break post-test assertions. We hand it
    a separate session bound to the same StaticPool engine so closing
    is harmless and the in-memory data stays visible.

    Returns the download Response. Raises if prepare returns non-202;
    in that case the caller should call /prepare directly and assert
    the error code instead of using this helper.
    """
    prepare = client.post(f"/api/projects/{project_id}/export/camtrap-dp/prepare")
    assert prepare.status_code == 202, prepare.text
    job_id = prepare.json()["job_id"]

    from sqlalchemy.orm import sessionmaker

    from app.workers import camtrap_export_worker
    from tests.conftest import _engine  # noqa: PLC2701 — shared in-memory engine

    worker_session_factory = sessionmaker(bind=_engine)

    def _fake_get_db():
        s = worker_session_factory()
        try:
            yield s
        finally:
            s.close()

    # Make sure the worker reads any rows the fixture committed before
    # we forked off this helper. Without it the fresh worker session
    # could miss rows that are still buffered on the fixture session.
    db.commit()

    with patch.object(camtrap_export_worker, "get_db", _fake_get_db):
        asyncio.run(camtrap_export_worker.process_camtrap_export_job(job_id))

    return client.get(
        f"/api/projects/{project_id}/export/camtrap-dp/download?job_id={job_id}"
    )

# ---------------------------------------------------------------------------
# Factory sugar
# ---------------------------------------------------------------------------


def _build_simple_project(db, *, timezone: str = "UTC", counting_threshold: float = 0.5):
    project = make_project(db, timezone=timezone, counting_threshold=counting_threshold)
    site = make_site(db, project_id=project.id, name="alpha", latitude=52.1, longitude=5.1)
    deployment = make_deployment(
        db,
        site_id=site.id,
        start_date_local=date(2024, 6, 1),
        end_date_local=date(2024, 6, 10),
        camera_model="Cam-X",
        camera_serial="SN-1",
    )
    db.commit()
    return project, site, deployment


# ---------------------------------------------------------------------------
# Observations
# ---------------------------------------------------------------------------


def test_export_detections_csv_happy_path(client, db):
    project, _site, deployment = _build_simple_project(db, timezone="Europe/Amsterdam")
    f_june = make_file(
        db,
        deployment_id=deployment.id,
        captured_at_local=datetime(2024, 6, 15, 9, 0, 0),
    )
    make_detection(db, file_id=f_june.id, category="animal", confidence=0.9, label="deer")
    make_detection(db, file_id=f_june.id, category="animal", confidence=0.7, label="deer")
    make_detection(db, file_id=f_june.id, category="person", confidence=0.95)
    f_dec = make_file(
        db,
        deployment_id=deployment.id,
        captured_at_local=datetime(2024, 12, 15, 3, 0, 0),
        observation_type="blank",
    )
    db.commit()

    resp = client.get(f"/api/projects/{project.id}/export/detections?format=csv")
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("text/csv")
    assert "attachment; filename=" in resp.headers["content-disposition"]

    rows = list(csv.reader(io.StringIO(resp.content.decode("utf-8"))))
    headers = rows[0]
    # Lean detections table: detection_id + file_id + the detection's own
    # fields. Time / place live in files.csv (join on file_id).
    assert headers == [
        "detection_id", "file_id", "deployment_id", "event_id",
        "detection_category", "detection_confidence",
        "classification_label", "classification_confidence",
        "ai_classification_label", "ai_classification_confidence",
        "classification_method", "is_verified",
        "taxon_class", "taxon_order", "taxon_family", "taxon_genus",
        "taxon_species", "taxon_variant", "scientific_name", "common_name",
        "frame_number", "bbox_x", "bbox_y", "bbox_width", "bbox_height",
    ]
    cls_i = headers.index("classification_label")
    cat_i = headers.index("detection_category")
    conf_i = headers.index("detection_confidence")

    data = rows[1:]
    # One row per real detection: 2 deer + 1 person. The empty December
    # file produces no detection row (it lives in the Files export).
    assert len(data) == 3
    assert all(r[cat_i] != "blank" for r in data)
    assert f_dec.id not in {r[headers.index("file_id")] for r in data}

    # Each deer detection keeps its own confidence (no max-aggregation).
    deer = [r for r in data if r[cls_i] == "deer"]
    assert len(deer) == 2
    assert sorted(float(r[conf_i]) for r in deer) == pytest.approx([0.7, 0.9], abs=1e-4)

    # Person: detected but not species-classified, so classification is empty.
    person = next(r for r in data if r[cat_i] == "person")
    assert float(person[conf_i]) == pytest.approx(0.95, abs=1e-4)
    assert person[cls_i] == ""


def test_export_detections_records_ai_vs_human_labels(client, db):
    """The detections CSV keeps the AI's original call alongside the
    current (possibly human-corrected) label, so a correction is visible."""
    project, _site, deployment = _build_simple_project(db)
    f = make_file(
        db, deployment_id=deployment.id, captured_at_local=datetime(2024, 6, 1, 9, 0, 0)
    )
    # AI said "wallaby"; a human later corrected it to "possum" and verified.
    make_detection(
        db,
        file_id=f.id,
        category="animal",
        confidence=0.9,
        label="possum",
        label_confidence=1.0,
        original_label="wallaby",
        original_label_confidence=0.42,
        classification_method="human",
        verified=True,
    )
    db.commit()

    resp = client.get(f"/api/projects/{project.id}/export/detections?format=csv")
    assert resp.status_code == 200
    rows = list(csv.reader(io.StringIO(resp.content.decode("utf-8"))))
    headers, data = rows[0], rows[1:]
    row = data[0]

    def cell(name: str) -> str:
        return row[headers.index(name)]

    assert cell("classification_label") == "possum"       # current
    assert cell("ai_classification_label") == "wallaby"    # AI's final call, retained
    assert float(cell("ai_classification_confidence")) == pytest.approx(0.42)
    assert cell("classification_method") == "human"
    assert cell("is_verified") == "TRUE"


def test_export_ai_label_matches_current_when_unverified(client, db):
    """For an untouched machine detection, ai_classification_label equals
    classification_label: both show the surfaced (post-rollup) call, so the
    export never exposes a label the user never saw."""
    project, _site, deployment = _build_simple_project(db)
    f = make_file(
        db, deployment_id=deployment.id, captured_at_local=datetime(2024, 6, 1, 9, 0, 0)
    )
    # Machine-final label mirrored into original_label (what postprocessing does).
    make_detection(
        db,
        file_id=f.id,
        category="animal",
        confidence=0.9,
        label="equidae",
        label_confidence=0.98,
        original_label="equidae",
        original_label_confidence=0.98,
        classification_method="machine",
        verified=False,
    )
    db.commit()

    resp = client.get(f"/api/projects/{project.id}/export/detections?format=csv")
    assert resp.status_code == 200
    rows = list(csv.reader(io.StringIO(resp.content.decode("utf-8"))))
    headers, data = rows[0], rows[1:]
    row = data[0]

    def cell(name: str) -> str:
        return row[headers.index(name)]

    assert cell("classification_label") == "equidae"
    assert cell("ai_classification_label") == "equidae"   # same as current
    assert cell("classification_method") == "machine"
    assert cell("is_verified") == "FALSE"


def test_export_scope_by_site_and_deployment(client, db):
    """Exports narrow to a site (all its deployments) or a specific
    deployment; no scope exports the whole project."""
    project = make_project(db)
    site_a = make_site(db, project_id=project.id, name="alpha")
    site_b = make_site(db, project_id=project.id, name="beta")
    dep_a = make_deployment(db, site_id=site_a.id)
    dep_b = make_deployment(db, site_id=site_b.id)
    f_a = make_file(
        db, deployment_id=dep_a.id, captured_at_local=datetime(2024, 6, 1, 9, 0, 0)
    )
    f_b = make_file(
        db, deployment_id=dep_b.id, captured_at_local=datetime(2024, 6, 2, 9, 0, 0)
    )
    make_detection(db, file_id=f_a.id, category="animal", confidence=0.9, label="deer")
    make_detection(db, file_id=f_b.id, category="animal", confidence=0.9, label="fox")
    db.commit()

    base = f"/api/projects/{project.id}/export/detections?format=csv"

    def _labels(url: str) -> set[str]:
        resp = client.get(url)
        assert resp.status_code == 200
        rows = list(csv.reader(io.StringIO(resp.content.decode("utf-8"))))
        i = rows[0].index("classification_label")
        return {r[i] for r in rows[1:]}

    # Whole project (no scope): both detections.
    assert _labels(base) == {"deer", "fox"}
    # Scope to site A → only its deployment's detection.
    assert _labels(f"{base}&site_ids={site_a.id}") == {"deer"}
    # Scope to deployment B → only that deployment's detection.
    assert _labels(f"{base}&deployment_ids={dep_b.id}") == {"fox"}
    # Site A + deployment B combined → union.
    assert _labels(
        f"{base}&site_ids={site_a.id}&deployment_ids={dep_b.id}"
    ) == {"deer", "fox"}

    # The deployments table scopes through its own query path too.
    dep_resp = client.get(
        f"/api/projects/{project.id}/export/deployments"
        f"?format=csv&site_ids={site_a.id}"
    )
    dep_rows = list(csv.reader(io.StringIO(dep_resp.content.decode("utf-8"))))
    dep_id_i = dep_rows[0].index("deployment_id")
    assert {r[dep_id_i] for r in dep_rows[1:]} == {dep_a.id}


def test_export_files_includes_empties(client, db):
    """The Files export lists every file once, including empties, which is
    where 'which files had no detections' lives (not faked into detections)."""
    project, _site, deployment = _build_simple_project(db, timezone="Europe/Amsterdam")
    f_animal = make_file(
        db,
        deployment_id=deployment.id,
        captured_at_local=datetime(2024, 6, 15, 9, 0, 0),
        observation_type="animal",
    )
    make_detection(db, file_id=f_animal.id, category="animal", confidence=0.9, label="deer")
    f_blank = make_file(
        db,
        deployment_id=deployment.id,
        captured_at_local=datetime(2024, 12, 15, 3, 0, 0),
        observation_type="blank",
    )
    db.commit()

    resp = client.get(f"/api/projects/{project.id}/export/files?format=csv")
    assert resp.status_code == 200
    rows = list(csv.reader(io.StringIO(resp.content.decode("utf-8"))))
    headers = rows[0]
    assert headers[0] == "file_id"
    assert "observation_type" in headers
    assert "event_id" in headers
    assert "file_type" in headers
    fid_i = headers.index("file_id")
    cat_i = headers.index("observation_type")
    dt_i = headers.index("datetime")

    data = rows[1:]
    by_id = {r[fid_i]: r for r in data}
    # Every row is as wide as the header, for the populated and the blank
    # file alike. The block describing the deciding box is built by two
    # branches with hand-kept widths, and a short row does not raise: CSV
    # writes it happily and every column after the gap silently shifts left.
    assert all(len(r) == len(headers) for r in data)
    # Both files appear, exactly once each; the empty file's observation_type
    # is blank.
    assert set(by_id) == {f_animal.id, f_blank.id}
    assert by_id[f_blank.id][cat_i] == "blank"
    assert by_id[f_animal.id][cat_i] == "animal"
    # DST-correct per-file offset (datetime lives on the files table now).
    assert by_id[f_animal.id][dt_i].endswith("+02:00")
    assert by_id[f_blank.id][dt_i].endswith("+01:00")


def test_files_export_camera_columns_come_from_stored_exif(client, db):
    """The four camera columns sit directly after datetime and read the EXIF
    block stored at analysis time (File.exif_data). A file without one, a
    video, an old analysis, or a camera that writes no such tags, reads
    four blanks."""
    project, _site, deployment = _build_simple_project(db)
    # NUL padding and trailing spaces are what real files carry: EXIF ASCII
    # fields are fixed-length, and the detector's reader keeps the padding
    # ('HC500 HYPERFIRE\x00\x00\x00\x00', 'Reconyx  ' — observed in the
    # example-data test images). Bare NULs in a cell crash openpyxl.
    f_exif = make_file(
        db,
        deployment_id=deployment.id,
        observation_type="blank",
        exif_data={
            "DateTimeOriginal": "2024:06:15 09:00:00",
            "Make": "RECONYX ",
            "Model": "HC600 HYPERFIRE\x00\x00\x00\x00",
            "AmbientTemperature": "23.65",
            "BodySerialNumber": "H600FF01234567",
        },
    )
    # A Make that is nothing but padding reads blank, not garbage.
    f_nul = make_file(
        db,
        deployment_id=deployment.id,
        observation_type="blank",
        exif_data={"Make": "\x00\x00\x00\x00\x00\x00\x00"},
    )
    # Only the date, the shape every pre-change analysis stored.
    f_dated = make_file(
        db,
        deployment_id=deployment.id,
        observation_type="blank",
        exif_data={"DateTimeOriginal": "2024:06:15 09:00:00"},
    )
    f_bare = make_file(db, deployment_id=deployment.id, observation_type="blank")
    db.commit()

    headers, by_id = _files_rows(client, project.id)
    dt_i = headers.index("datetime")
    assert headers[dt_i : dt_i + 5] == [
        "datetime",
        "camera_make",
        "camera_model",
        "ambient_temperature",
        "camera_serial",
    ]

    def camera_cells(row: list[str]) -> list[str]:
        return row[dt_i + 1 : dt_i + 5]

    assert camera_cells(by_id[f_exif.id]) == [
        "RECONYX",
        "HC600 HYPERFIRE",
        "23.65",
        "H600FF01234567",
    ]
    assert camera_cells(by_id[f_nul.id]) == ["", "", "", ""]
    assert camera_cells(by_id[f_dated.id]) == ["", "", "", ""]
    assert camera_cells(by_id[f_bare.id]) == ["", "", "", ""]


# ---------------------------------------------------------------------------
# The one label a file gets
#
# observation_type plus the species block beside it all describe the SAME box:
# the file's strongest passing detection. These tests pin that they never come
# from different boxes.
# ---------------------------------------------------------------------------


def _files_rows(client, project_id: str):
    """`(headers, {file_id: row})` for the Files CSV export."""
    resp = client.get(f"/api/projects/{project_id}/export/files?format=csv")
    assert resp.status_code == 200, resp.text
    rows = list(csv.reader(io.StringIO(resp.content.decode("utf-8"))))
    headers = rows[0]
    fid_i = headers.index("file_id")
    return headers, {r[fid_i]: r for r in rows[1:]}


_RANK_COLUMNS = (
    "taxon_class",
    "taxon_order",
    "taxon_family",
    "taxon_genus",
    "taxon_species",
)


def _species(headers: list[str], row: list[str]) -> list[str]:
    """The label and the two display names of a Files row. The ranks are
    checked by _ranks, so the common assertions stay readable."""
    return [
        row[headers.index(name)]
        for name in ("classification_label", "scientific_name", "common_name")
    ]


def _ranks(headers: list[str], row: list[str]) -> list[str]:
    """The five formal ranks of a Files row, broad to specific."""
    return [row[headers.index(name)] for name in _RANK_COLUMNS]


def _confidences(headers: list[str], row: list[str]) -> list[str]:
    """`[detection_confidence, classification_confidence]` of a Files row."""
    return [
        row[headers.index(name)]
        for name in ("detection_confidence", "classification_confidence")
    ]


def test_files_export_names_the_strongest_detections_species(client, db):
    """The Files table answers "what is this file" on its own: the category of
    the strongest box, then that same box's species, ranks and display names."""
    project, _site, deployment = _build_simple_project(db)
    taxonomy = LabelTaxonomy(
        classification_model_id="",
        project_id=project.id,
        name="red fox",
        level="species",
        taxon_class="mammalia",
        taxon_order="carnivora",
        taxon_family="canidae",
        taxon_genus="vulpes",
        taxon_species="vulpes",
    )
    db.add(taxonomy)
    db.flush()
    f = make_file(db, deployment_id=deployment.id, observation_type="animal")
    make_detection(
        db,
        file_id=f.id,
        category="animal",
        confidence=0.9,
        label="red fox",
        label_confidence=0.81,
        label_taxonomy_id=taxonomy.id,
        scientific_name="Vulpes vulpes",
        common_name="Red fox",
    )
    db.commit()

    headers, by_id = _files_rows(client, project.id)
    # Every score sits directly after what it scores, the way detections.csv
    # lays these two out. Pinning the pairing rather than mere contiguity:
    # a bare confidence eight columns from its subject is unreadable, and
    # two of them side by side are indistinguishable.
    obs_i = headers.index("observation_type")
    assert headers[obs_i : obs_i + 12] == [
        "observation_type",
        "detection_confidence",
        "classification_label",
        "classification_confidence",
        "taxon_class",
        "taxon_order",
        "taxon_family",
        "taxon_genus",
        "taxon_species",
        "taxon_variant",
        "scientific_name",
        "common_name",
    ]
    assert headers[0] == "file_id"
    assert _species(headers, by_id[f.id]) == ["red fox", "Vulpes vulpes", "Red fox"]
    assert _confidences(headers, by_id[f.id]) == ["0.9", "0.81"]
    assert _ranks(headers, by_id[f.id]) == [
        "mammalia",
        "carnivora",
        "canidae",
        "vulpes",
        "vulpes",
    ]


def test_files_export_ranks_say_which_rank_the_label_is(client, db):
    """Taxonomic rollup puts species, families and orders in one label column,
    so "rodentia" and "porcupine" sit side by side. The ranks are the only
    thing that says which is which; without them, grouping by the label
    silently merges an order with the species inside it."""
    project, _site, deployment = _build_simple_project(db)
    order_level = LabelTaxonomy(
        classification_model_id="",
        project_id=project.id,
        name="rodentia",
        level="order",
        taxon_class="mammalia",
        taxon_order="rodentia",
    )
    db.add(order_level)
    db.flush()
    f = make_file(db, deployment_id=deployment.id, observation_type="animal")
    make_detection(
        db,
        file_id=f.id,
        category="animal",
        confidence=0.9,
        label="rodentia",
        label_taxonomy_id=order_level.id,
        scientific_name="Rodentia",
        common_name="Rodentia",
    )
    db.commit()

    headers, by_id = _files_rows(client, project.id)
    # Filled down to order, empty below it: this label is not a species.
    assert _ranks(headers, by_id[f.id]) == ["mammalia", "rodentia", "", "", ""]


def test_files_export_confidences_come_from_the_deciding_box(client, db):
    """Both numbers must be the winning box's own, not a maximum taken over
    the file and not each other.

    Every value is rank-inverted so each plausible wrong source yields a
    different number. The winner has neither the highest detector score nor
    the highest label score, and its own two scores run opposite ways, so a
    max over the file, a read off the best-labelled box, or the two columns
    swapped all produce something visibly wrong."""
    project, _site, deployment = _build_simple_project(db)
    f = make_file(db, deployment_id=deployment.id, observation_type="animal")
    # Wins on the verified tier alone.
    make_detection(
        db,
        file_id=f.id,
        category="animal",
        confidence=0.80,
        verified=True,
        label="red fox",
        label_confidence=0.11,
        scientific_name="Vulpes vulpes",
        common_name="Red fox",
    )
    # Higher on both scores, but unverified.
    make_detection(
        db,
        file_id=f.id,
        category="animal",
        confidence=0.99,
        label="deer",
        label_confidence=0.98,
        scientific_name="Cervidae",
        common_name="Deer",
    )
    make_detection(
        db,
        file_id=f.id,
        category="animal",
        confidence=0.60,
        label="badger",
        label_confidence=0.55,
        scientific_name="Meles meles",
        common_name="Badger",
    )
    db.commit()

    headers, by_id = _files_rows(client, project.id)
    row = by_id[f.id]
    assert _species(headers, row) == ["red fox", "Vulpes vulpes", "Red fox"]
    assert _confidences(headers, row) == ["0.8", "0.11"]
    # Nothing from the decoy leaked into a column this test does not name.
    assert "0.99" not in row
    assert "0.98" not in row
    assert "deer" not in row


def test_files_export_does_not_take_the_best_label_off_a_weaker_box(client, db):
    """The camouflage regression, and the reason this column is defined the way
    it is. A clip of a person inspecting a camera produced person boxes at 0.65
    to 0.95 plus one false-positive animal box the classifier called chimpanzee
    at 29%. Reporting the best *label* labels that file a chimpanzee; reporting
    the strongest *box* calls it a person, which is what the picture shows."""
    project, _site, deployment = _build_simple_project(db)
    # Give the false positive a full taxonomy, so the assertions below prove
    # the ranks were not taken from it rather than merely finding it empty.
    ape = LabelTaxonomy(
        classification_model_id="",
        project_id=project.id,
        name="chimpanzee",
        level="species",
        taxon_class="mammalia",
        taxon_order="primates",
        taxon_family="hominidae",
        taxon_genus="pan",
        taxon_species="troglodytes",
    )
    db.add(ape)
    db.flush()
    f = make_file(db, deployment_id=deployment.id, observation_type="person")
    make_detection(
        db,
        file_id=f.id,
        category="person",
        confidence=0.95,
        scientific_name="Person",
        common_name="Person",
    )
    make_detection(
        db,
        file_id=f.id,
        category="animal",
        confidence=0.677,
        label="chimpanzee",
        label_taxonomy_id=ape.id,
        scientific_name="Pan troglodytes",
        common_name="Chimpanzee",
    )
    db.commit()

    headers, by_id = _files_rows(client, project.id)
    assert _species(headers, by_id[f.id]) == ["", "Person", "Person"]
    # The chimpanzee's taxonomy must not leak in either: the ranks describe
    # the winning box, which is a person and has none.
    assert _ranks(headers, by_id[f.id]) == ["", "", "", "", ""]
    assert "chimpanzee" not in by_id[f.id]
    assert "primates" not in by_id[f.id]


def test_files_export_species_are_blank_when_no_box_passes(client, db):
    """No trusted box means there is nothing to name. observation_type already
    carries "blank"; inventing a name here would put a file state into a
    species column and break the join to detections.csv."""
    project, _site, deployment = _build_simple_project(db)
    f_empty = make_file(db, deployment_id=deployment.id, observation_type="blank")
    f_weak = make_file(db, deployment_id=deployment.id, observation_type="blank")
    make_detection(
        db,
        file_id=f_weak.id,
        category="animal",
        confidence=0.2,
        label="deer",
        scientific_name="Cervidae",
        common_name="Deer",
    )
    db.commit()

    headers, by_id = _files_rows(client, project.id)
    for file_id in (f_empty.id, f_weak.id):
        assert _species(headers, by_id[file_id]) == ["", "", ""]
        assert _ranks(headers, by_id[file_id]) == ["", "", "", "", ""]
        # Blank, never "0": a zero would read as a real measurement of zero
        # and would survive a `< x` filter.
        assert _confidences(headers, by_id[file_id]) == ["", ""]
        assert by_id[file_id][headers.index("observation_type")] == "blank"


def test_files_export_follows_the_verified_box(client, db):
    """A human looked at the fox box, so it outranks a model that is merely
    more confident about a deer. Same ordering the folder tree uses."""
    project, _site, deployment = _build_simple_project(db)
    f = make_file(db, deployment_id=deployment.id, observation_type="animal")
    make_detection(
        db,
        file_id=f.id,
        category="animal",
        confidence=0.99,
        label="deer",
        scientific_name="Cervidae",
        common_name="Deer",
    )
    make_detection(
        db,
        file_id=f.id,
        category="animal",
        confidence=0.30,
        verified=True,
        label="red fox",
        scientific_name="Vulpes vulpes",
        common_name="Red fox",
    )
    db.commit()

    headers, by_id = _files_rows(client, project.id)
    assert _species(headers, by_id[f.id]) == ["red fox", "Vulpes vulpes", "Red fox"]


def test_video_exports_describe_only_the_visible_frame(client, db):
    """One row stands for one video, so it describes the one frame that
    stands for that video. Here the strongest box overall is a deer on
    frame 7, which was never written to disk; the frame the user can open
    holds a person. The row reports the person, and the deer is absent
    from the detections table too, because a box on a frame nobody can
    open cannot be seen, filtered to or relabelled. It survives in
    addaxai-recognitions.json, which is the complete record."""
    project, _site, deployment = _build_simple_project(db)
    f = make_file(
        db,
        deployment_id=deployment.id,
        file_type="video",
        file_format="mp4",
        best_frame_number=3,
        observation_type="person",
    )
    make_detection(
        db,
        file_id=f.id,
        category="animal",
        confidence=0.9,
        frame_number=7,
        label="deer",
        scientific_name="Cervidae",
        common_name="Deer",
    )
    make_detection(
        db,
        file_id=f.id,
        category="person",
        confidence=0.6,
        frame_number=3,
        scientific_name="Person",
        common_name="Person",
    )
    db.commit()

    headers, by_id = _files_rows(client, project.id)
    assert _species(headers, by_id[f.id]) == ["", "Person", "Person"]
    # 0.6 is the person on the saved frame, not the 0.9 deer on frame 7, so
    # the confidence respects visible_detections and not only the strongest
    # box overall. Empty label score: a person carries no species.
    assert _confidences(headers, by_id[f.id]) == ["0.6", ""]

    # The per-box grain agrees: the off-frame deer is not offered as a
    # species the user could go and correct, because there is no picture
    # of it anywhere in the app. Only the person on the saved frame is.
    resp = client.get(f"/api/projects/{project.id}/export/detections?format=csv")
    det_rows = list(csv.reader(io.StringIO(resp.content.decode("utf-8"))))
    det_headers = det_rows[0]
    frames = {
        r[det_headers.index("frame_number")]
        for r in det_rows[1:]
        if r[det_headers.index("file_id")] == f.id
    }
    assert frames == {"3"}
    labels = {
        r[det_headers.index("classification_label")]
        for r in det_rows[1:]
        if r[det_headers.index("file_id")] == f.id
    }
    assert "deer" not in labels


def test_files_export_video_with_empty_best_frame_reads_blank(client, db):
    """A video whose best frame holds nothing above the threshold reports
    blank, even though a confident box exists on another frame. That box has
    no card in the Labels grid, no MaxN count and no crop, so blank is the
    truthful summary of what the user can see and act on. No fallback to the
    other frames: that would be a second rule firing only sometimes."""
    project, _site, deployment = _build_simple_project(db)
    f = make_file(
        db,
        deployment_id=deployment.id,
        file_type="video",
        file_format="mp4",
        best_frame_number=3,
        observation_type="blank",
    )
    make_detection(
        db,
        file_id=f.id,
        category="animal",
        confidence=0.9,
        frame_number=7,
        label="deer",
        scientific_name="Cervidae",
        common_name="Deer",
    )
    db.commit()

    headers, by_id = _files_rows(client, project.id)
    assert _species(headers, by_id[f.id]) == ["", "", ""]
    assert _ranks(headers, by_id[f.id]) == ["", "", "", "", ""]


def test_files_export_verified_box_passes_on_any_frame(client, db):
    """The escape hatch. A human decision must never be out of reach, so a
    verified box counts wherever it sits, even though its thumbnail is
    missing."""
    project, _site, deployment = _build_simple_project(db)
    f = make_file(
        db,
        deployment_id=deployment.id,
        file_type="video",
        file_format="mp4",
        best_frame_number=3,
        observation_type="animal",
    )
    make_detection(
        db,
        file_id=f.id,
        category="animal",
        confidence=0.30,
        frame_number=7,
        verified=True,
        label="red fox",
        scientific_name="Vulpes vulpes",
        common_name="Red fox",
    )
    make_detection(
        db,
        file_id=f.id,
        category="person",
        confidence=0.6,
        frame_number=3,
        scientific_name="Person",
        common_name="Person",
    )
    db.commit()

    headers, by_id = _files_rows(client, project.id)
    assert _species(headers, by_id[f.id]) == ["red fox", "Vulpes vulpes", "Red fox"]


def test_files_export_species_follow_the_export_scope(client, db):
    """excluded_classes drops boxes from the export, but File.observation_type
    is a stored column derived without exclusions, so the two can describe
    different boxes when the exclusion list changed after analysis and nothing
    reprocessed. That is deliberate: re-deriving observation_type here would
    silently change a shipped column and make the export disagree with the app.
    The species columns keep the more useful promise instead, that a non-empty
    value always resolves to a row in detections.csv under the same file_id."""
    project, _site, deployment = _build_simple_project(db)
    project.excluded_classes = ["dog"]
    f = make_file(db, deployment_id=deployment.id, observation_type="animal")
    make_detection(
        db,
        file_id=f.id,
        category="animal",
        confidence=0.95,
        label="dog",
        scientific_name="Canis familiaris",
        common_name="Dog",
    )
    make_detection(
        db,
        file_id=f.id,
        category="person",
        confidence=0.6,
        scientific_name="Person",
        common_name="Person",
    )
    db.commit()

    headers, by_id = _files_rows(client, project.id)
    assert by_id[f.id][headers.index("observation_type")] == "animal"
    assert _species(headers, by_id[f.id]) == ["", "Person", "Person"]


def test_event_id_is_blank_when_a_file_has_no_event(client, db):
    """`event_id` must never stand in a file id for a missing event: a
    consumer joining files.csv to counts.csv on event_id would match
    nothing and never know why. Blank says "no event", which is true.

    In practice every image / video is clustered into exactly one event
    (date-less files become singleton events), so this only fires when
    events have not been generated yet.
    """
    project, _site, deployment = _build_simple_project(db)
    f_no_event = make_file(
        db,
        deployment_id=deployment.id,
        captured_at_local=datetime(2024, 6, 1, 12, 0, 0),
        observation_type="animal",
    )
    db.commit()

    resp = client.get(f"/api/projects/{project.id}/export/files?format=csv")
    assert resp.status_code == 200
    rows = list(csv.reader(io.StringIO(resp.content.decode("utf-8"))))
    headers = rows[0]
    fid_i = headers.index("file_id")
    eid_i = headers.index("event_id")

    row = next(r for r in rows[1:] if r[fid_i] == f_no_event.id)
    assert row[eid_i] == ""
    assert row[eid_i] != f_no_event.id


def test_detections_event_id_is_blank_when_a_file_has_no_event(client, db):
    """Same contract as the files export: the detections table must not
    stand a file id in for a missing event either."""
    project, _site, deployment = _build_simple_project(db)
    f_no_event = make_file(
        db,
        deployment_id=deployment.id,
        captured_at_local=datetime(2024, 6, 1, 12, 0, 0),
        observation_type="animal",
    )
    make_detection(db, file_id=f_no_event.id, confidence=0.9)
    db.commit()

    resp = client.get(f"/api/projects/{project.id}/export/detections?format=csv")
    assert resp.status_code == 200
    rows = list(csv.reader(io.StringIO(resp.content.decode("utf-8"))))
    headers = rows[0]
    fid_i = headers.index("file_id")
    eid_i = headers.index("event_id")

    row = next(r for r in rows[1:] if r[fid_i] == f_no_event.id)
    assert row[eid_i] == ""
    assert row[eid_i] != f_no_event.id


def test_event_id_carries_the_real_event_id_when_clustered(client, db):
    """The other half of the contract: a clustered file reports its
    actual event id, in both the files and the detections table, so a
    join against counts.csv resolves."""
    project, _site, deployment = _build_simple_project(db)
    event = make_event_with_files(
        db,
        deployment_id=deployment.id,
        event_start_local=datetime(2024, 6, 1, 12, 0, 0),
    )
    clustered = (
        db.query(File).filter(File.deployment_id == deployment.id).one()
    )
    make_detection(db, file_id=clustered.id, confidence=0.9)
    db.commit()

    files_resp = client.get(f"/api/projects/{project.id}/export/files?format=csv")
    files_rows = list(csv.reader(io.StringIO(files_resp.content.decode("utf-8"))))
    fh = files_rows[0]
    frow = next(
        r for r in files_rows[1:] if r[fh.index("file_id")] == clustered.id
    )
    assert frow[fh.index("event_id")] == event.id

    det_resp = client.get(
        f"/api/projects/{project.id}/export/detections?format=csv"
    )
    det_rows = list(csv.reader(io.StringIO(det_resp.content.decode("utf-8"))))
    dh = det_rows[0]
    drow = next(
        r for r in det_rows[1:] if r[dh.index("file_id")] == clustered.id
    )
    assert drow[dh.index("event_id")] == event.id


def test_export_detections_tsv_and_xlsx(client, db):
    project, _site, deployment = _build_simple_project(db)
    f = make_file(db, deployment_id=deployment.id)
    make_detection(db, file_id=f.id, category="animal", confidence=0.8, label="fox")
    db.commit()

    resp_tsv = client.get(f"/api/projects/{project.id}/export/detections?format=tsv")
    assert resp_tsv.status_code == 200
    assert resp_tsv.headers["content-type"].startswith("text/tab-separated-values")
    tsv_rows = list(csv.reader(io.StringIO(resp_tsv.content.decode("utf-8")), delimiter="\t"))
    assert tsv_rows[0][0] == "detection_id"
    assert any("fox" in r for r in tsv_rows[1:])

    resp_xlsx = client.get(f"/api/projects/{project.id}/export/detections?format=xlsx")
    assert resp_xlsx.status_code == 200
    assert "spreadsheetml" in resp_xlsx.headers["content-type"]
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(resp_xlsx.content))
    ws = wb.active
    assert ws.title == "Detections"
    sheet_rows = list(ws.iter_rows(values_only=True))
    assert sheet_rows[0][0] == "detection_id"
    assert any(
        isinstance(v, str) and "fox" in v
        for row in sheet_rows[1:]
        for v in row
    )


def test_export_detections_respects_threshold_and_verified_override(client, db):
    project, _site, deployment = _build_simple_project(db, counting_threshold=0.5)
    f = make_file(db, deployment_id=deployment.id)
    # Below threshold, unverified → excluded.
    make_detection(db, file_id=f.id, category="animal", confidence=0.3, label="fox")
    # Below threshold, verified → included via override.
    make_detection(
        db, file_id=f.id, category="animal", confidence=0.2, label="deer", verified=True
    )
    # Above threshold → included.
    make_detection(db, file_id=f.id, category="animal", confidence=0.8, label="bear")
    db.commit()

    resp = client.get(f"/api/projects/{project.id}/export/detections?format=csv")
    text = resp.content.decode("utf-8")
    assert "deer" in text
    assert "bear" in text
    assert "fox" not in text


def test_export_detections_respects_excluded_classes(client, db):
    project = make_project(db, timezone="UTC", excluded_classes=["domestic_cat"])
    site = make_site(db, project_id=project.id)
    deployment = make_deployment(db, site_id=site.id)
    f = make_file(db, deployment_id=deployment.id)
    make_detection(db, file_id=f.id, category="animal", confidence=0.9, label="domestic_cat")
    make_detection(db, file_id=f.id, category="animal", confidence=0.9, label="fox")
    db.commit()

    resp = client.get(f"/api/projects/{project.id}/export/detections?format=csv")
    text = resp.content.decode("utf-8")
    assert "fox" in text
    assert "domestic_cat" not in text


def test_export_detections_verified_survives_excluded_classes(client, db):
    # A human relabel to an excluded species (possible when the species
    # selection hid the true class from the classifier) must not be
    # dropped from the export: verified outranks the exclusion config.
    project = make_project(db, timezone="UTC", excluded_classes=["bird"])
    site = make_site(db, project_id=project.id)
    deployment = make_deployment(db, site_id=site.id)
    f = make_file(db, deployment_id=deployment.id)
    make_detection(db, file_id=f.id, category="animal", confidence=0.9, label="coyote")
    # AI said coyote, human verified it as bird.
    make_detection(
        db, file_id=f.id, category="animal", confidence=0.9, label="bird", verified=True
    )
    # Excluded and unverified → still dropped.
    make_detection(db, file_id=f.id, category="animal", confidence=0.9, label="bird")
    db.commit()

    resp = client.get(f"/api/projects/{project.id}/export/detections?format=csv")
    rows = list(csv.reader(io.StringIO(resp.content.decode("utf-8"))))
    labels = [r[rows[0].index("classification_label")] for r in rows[1:]]
    assert "coyote" in labels
    assert labels.count("bird") == 1


def test_export_detections_project_not_found(client):
    resp = client.get("/api/projects/does-not-exist/export/detections?format=csv")
    assert resp.status_code == 404


def test_export_observations_event_level(client, db):
    """Event-level Observations: one row per species per event with the
    effective (human-confirmed, else AI) count."""
    from app.api.crud.event_observation import (
        calculate_max_n_for_event,
        set_human_count,
    )

    project, _site, deployment = _build_simple_project(db, timezone="UTC")
    ev = make_event_with_files(
        db,
        deployment_id=deployment.id,
        event_start_local=datetime(2024, 6, 15, 9, 0, 0),
    )
    # Two deer on the same frame → AI MaxN of 2 for the event.
    make_detection(db, file_id=ev.files[0].id, category="animal", confidence=0.9, label="deer")
    make_detection(db, file_id=ev.files[0].id, category="animal", confidence=0.8, label="deer")
    obs = calculate_max_n_for_event(db, ev.id, project.counting_threshold)
    # The rows above are still pending; `set_human_count` looks its row up
    # by id in SQL and silently returns None when it is not there yet.
    db.flush()
    # Human bumps the deer count above the per-frame max.
    set_human_count(db, obs[0].id, 5)
    db.commit()

    resp = client.get(f"/api/projects/{project.id}/export/observations?format=csv")
    assert resp.status_code == 200
    rows = list(csv.reader(io.StringIO(resp.content.decode("utf-8"))))
    headers = rows[0]
    assert headers[0] == "event_id"
    assert "count" in headers
    assert "is_confirmed" in headers
    assert "classification_label" in headers
    count_i = headers.index("count")
    label_i = headers.index("classification_label")

    data = rows[1:]
    deer = [r for r in data if r[label_i] == "deer"]
    # One event-level row for deer, carrying the human count (5), not 2.
    assert len(deer) == 1
    assert deer[0][count_i] == "5"


def test_export_observations_project_not_found(client):
    resp = client.get("/api/projects/does-not-exist/export/observations?format=csv")
    assert resp.status_code == 404


def test_export_spreadsheet_is_multi_sheet_workbook(client, db):
    """The combined Spreadsheet export is one XLSX with Counts, Detections,
    Files and Deployments sheets.

    Order is asserted, not just membership: a workbook opens on its first
    sheet, so that sheet is what a user takes the file to contain. Counts
    leads because it is the analysis-ready table the docs send people to
    first."""
    from openpyxl import load_workbook

    from app.api.crud.event_observation import calculate_max_n_for_event

    project, _site, deployment = _build_simple_project(db, timezone="UTC")
    ev = make_event_with_files(
        db,
        deployment_id=deployment.id,
        event_start_local=datetime(2024, 6, 15, 9, 0, 0),
    )
    make_detection(db, file_id=ev.files[0].id, category="animal", confidence=0.9, label="deer")
    calculate_max_n_for_event(db, ev.id, project.counting_threshold)
    db.commit()

    resp = client.get(f"/api/projects/{project.id}/export/spreadsheet")
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]

    wb = load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames == ["Counts", "Detections", "Files", "Deployments"]
    deployments = list(wb["Deployments"].iter_rows(values_only=True))
    files = list(wb["Files"].iter_rows(values_only=True))
    det = list(wb["Detections"].iter_rows(values_only=True))
    counts = list(wb["Counts"].iter_rows(values_only=True))
    assert deployments[0][0] == "deployment_id"
    assert files[0][0] == "file_id"
    assert det[0][0] == "detection_id"
    assert counts[0][0] == "event_id"
    # The deer appears as a detection row and an event-level count row.
    assert any("deer" in str(v) for row in det[1:] for v in row)
    assert any("deer" in str(v) for row in counts[1:] for v in row)


# ---------------------------------------------------------------------------
# XLSX row limit
# ---------------------------------------------------------------------------


def test_xlsx_accepts_a_sheet_filled_to_the_row_limit():
    """The header counts as a row, so the last accepted workbook holds
    XLSX_MAX_ROWS - 1 data rows. Pinning the boundary from below is what
    stops a future off-by-one from refusing exports that fit."""
    rows = [[i] for i in range(export_formats.XLSX_MAX_ROWS - 1)]
    # Only the counting is under test; building the real workbook here
    # would cost a gigabyte of XML for one assertion.
    export_formats._check_xlsx_row_limit([("Detections", ["n"], rows)])


def test_xlsx_refuses_one_row_past_the_limit():
    """openpyxl does not check this itself: in write-only mode it accepts
    any number of rows and saves a file whose row indexes run past the
    cap, which Excel then refuses to open. Without this guard the user
    gets a corrupt download and no explanation."""
    rows = [[i] for i in range(export_formats.XLSX_MAX_ROWS)]
    with pytest.raises(export_formats.XlsxRowLimitError) as excinfo:
        export_formats._check_xlsx_row_limit([("Detections", ["n"], rows)])

    message = str(excinfo.value)
    # The message is shown to the user verbatim, as the 422 detail and as
    # a folder-run module error, so it has to name the table, its size
    # and the way out.
    assert "detections table" in message
    assert f"{export_formats.XLSX_MAX_ROWS:,}" in message
    assert "CSV" in message


def test_xlsx_row_limit_is_reported_as_422_not_a_corrupt_download(client, db):
    """The export endpoints turn the refusal into a 422 whose detail is
    the message itself; the frontend surfaces a string detail verbatim."""
    project, _site, deployment = _build_simple_project(db)
    file = make_file(db, deployment_id=deployment.id)
    make_detection(db, file_id=file.id, category="animal", confidence=0.9, label="deer")
    db.commit()

    # One real detection, and a limit low enough that it exceeds it. Far
    # cheaper than materialising a million rows, and it exercises the
    # same path the real overflow takes.
    with patch.object(export_formats, "XLSX_MAX_ROWS", 1):
        resp = client.get(
            f"/api/projects/{project.id}/export/detections?format=xlsx"
        )
    assert resp.status_code == 422
    assert "CSV" in resp.json()["detail"]

    # CSV has no such limit, so the same data still comes out.
    resp = client.get(f"/api/projects/{project.id}/export/detections?format=csv")
    assert resp.status_code == 200


# ---------------------------------------------------------------------------
# Spatial
# ---------------------------------------------------------------------------


def test_export_spatial_geojson(client, db):
    project, site, deployment = _build_simple_project(db)
    taxonomy = LabelTaxonomy(
        id=str(uuid.uuid4()),
        classification_model_id="TEST-MODEL",
        name="fox",
        scientific_name="Vulpes vulpes",
        level="species",
        taxon_class="mammalia",
        taxon_order="carnivora",
        taxon_family="canidae",
        taxon_genus="vulpes",
        taxon_species="vulpes",
    )
    db.add(taxonomy)
    db.flush()
    f = make_file(db, deployment_id=deployment.id)
    make_detection(
        db, file_id=f.id, category="animal", confidence=0.9,
        label="fox", label_taxonomy_id=taxonomy.id,
    )
    db.commit()

    resp = client.get(f"/api/projects/{project.id}/export/spatial?format=geojson")
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("application/geo+json")
    payload = json.loads(resp.content)
    assert payload["type"] == "FeatureCollection"

    # Two genuinely spatial layers only; the per-detection points are gone.
    layers = {feat["properties"]["layer"] for feat in payload["features"]}
    assert layers == {"deployments", "species_summary"}

    summary = next(
        feat for feat in payload["features"]
        if feat["properties"]["layer"] == "species_summary"
    )
    props = summary["properties"]
    assert props["classification_label"] == "fox"
    assert props["scientific_name"] == "Vulpes vulpes"
    assert props["taxon_genus"] == "vulpes"
    assert props["taxon_species"] == "vulpes"
    assert props["total_count"] == 1

    for feat in payload["features"]:
        assert feat["geometry"]["type"] == "Point"
        lon, lat = feat["geometry"]["coordinates"]
        assert lon == site.longitude
        assert lat == site.latitude


def test_export_spatial_shapefile_zip(client, db):
    project, _site, deployment = _build_simple_project(db)
    f = make_file(db, deployment_id=deployment.id)
    make_detection(db, file_id=f.id, category="animal", confidence=0.9, label="fox")
    db.commit()

    resp = client.get(f"/api/projects/{project.id}/export/spatial?format=shapefile")
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("application/zip")

    with zipfile.ZipFile(io.BytesIO(resp.content)) as zf:
        names = set(zf.namelist())
    expected = {
        "deployments.shp", "deployments.shx", "deployments.dbf", "deployments.prj",
        "species_summary.shp", "species_summary.shx", "species_summary.dbf",
        "species_summary.prj",
    }
    assert expected.issubset(names)
    assert not any(n.startswith("observations.") for n in names)


def test_export_spatial_gpkg(client, db):
    project, _site, deployment = _build_simple_project(db)
    taxonomy = LabelTaxonomy(
        id=str(uuid.uuid4()),
        classification_model_id="TEST-MODEL",
        name="fox",
        scientific_name="Vulpes vulpes",
        level="species",
        taxon_genus="vulpes",
        taxon_species="vulpes",
    )
    db.add(taxonomy)
    db.flush()
    f = make_file(db, deployment_id=deployment.id)
    make_detection(
        db, file_id=f.id, category="animal", confidence=0.9,
        label="fox", label_taxonomy_id=taxonomy.id,
    )
    db.commit()

    resp = client.get(f"/api/projects/{project.id}/export/spatial?format=gpkg")
    assert resp.status_code == 200

    # Round-trip through sqlite3 to confirm the feature tables and that the
    # species_summary attributes are actually populated (not silently blank).
    with tempfile.NamedTemporaryFile(suffix=".gpkg", delete=False) as tmp:
        tmp.write(resp.content)
        tmp_path = tmp.name
    try:
        conn = sqlite3.connect(tmp_path)
        tables = {
            row[0]
            for row in conn.execute(
                "SELECT table_name FROM gpkg_contents ORDER BY table_name"
            )
        }
        row = conn.execute(
            "SELECT classification_label, scientific_name, taxon_species, "
            "total_count FROM species_summary"
        ).fetchone()
        conn.close()
    finally:
        import os

        os.unlink(tmp_path)
    assert tables == {"deployments", "species_summary"}
    assert row == ("fox", "Vulpes vulpes", "vulpes", 1)


# ---------------------------------------------------------------------------
# CamTrap DP
# ---------------------------------------------------------------------------


def test_export_camtrap_dp_happy_path(client, db):
    project, _site, deployment = _build_simple_project(db, timezone="Europe/Amsterdam")
    # Link detection to a taxonomy row for scientificName.
    taxonomy = LabelTaxonomy(
        id=str(uuid.uuid4()),
        classification_model_id="TEST-MODEL",
        name="fox",
        scientific_name="Vulpes vulpes",
        level="species",
    )
    db.add(taxonomy)
    db.flush()

    f = make_file(
        db,
        deployment_id=deployment.id,
        captured_at_local=datetime(2024, 6, 15, 9, 0, 0),
    )
    make_detection(
        db,
        file_id=f.id,
        category="animal",
        confidence=0.9,
        label="fox",
        scientific_name="fox",
        label_confidence=0.88,
        label_taxonomy_id=taxonomy.id,
    )
    db.commit()

    resp = _run_camtrap_dp_export(client, db, project.id)
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("application/zip")

    with zipfile.ZipFile(io.BytesIO(resp.content)) as zf:
        names = set(zf.namelist())
        assert names == {
            "datapackage.json",
            "deployments.csv",
            "media.csv",
            "observations.csv",
        }
        dp = json.loads(zf.read("datapackage.json"))
        deps_rows = list(csv.reader(io.StringIO(zf.read("deployments.csv").decode())))
        media_rows = list(csv.reader(io.StringIO(zf.read("media.csv").decode())))
        obs_rows = list(csv.reader(io.StringIO(zf.read("observations.csv").decode())))

    assert dp["title"] == project.name
    assert dp["name"].startswith("addaxai-")
    assert dp["temporal"]["start"] == "2024-06-15"
    assert any(
        entry["scientificName"] == "Vulpes vulpes" for entry in dp["taxonomic"]
    )

    # CSV header sanity.
    assert deps_rows[0][0] == "deploymentID"
    assert media_rows[0] == [
        "mediaID", "deploymentID", "captureMethod", "timestamp",
        "filePath", "filePublic", "fileName", "fileMediatype",
        "exifData", "favorite", "mediaComments",
    ]
    assert obs_rows[0][0] == "observationID"
    # One detection → one animal observation row.
    assert len(obs_rows) == 2
    assert obs_rows[1][7] == "animal"
    assert obs_rows[1][9] == "Vulpes vulpes"


def test_export_camtrap_dp_422_when_no_deployments(client, db):
    project = make_project(db, timezone="UTC")
    db.commit()
    # 422 fires inside /prepare before the worker is dispatched, so we
    # bypass the helper and inspect the prepare response directly.
    resp = client.post(f"/api/projects/{project.id}/export/camtrap-dp/prepare")
    assert resp.status_code == 422


def test_export_camtrap_dp_blank_row_for_file_without_detections(client, db):
    project, _site, deployment = _build_simple_project(db)
    f = make_file(
        db,
        deployment_id=deployment.id,
        captured_at_local=datetime(2024, 6, 15, 9, 0, 0),
        observation_type="blank",
    )
    db.commit()

    resp = _run_camtrap_dp_export(client, db, project.id)
    assert resp.status_code == 200
    with zipfile.ZipFile(io.BytesIO(resp.content)) as zf:
        obs_rows = list(csv.reader(io.StringIO(zf.read("observations.csv").decode())))
    assert len(obs_rows) == 2
    assert obs_rows[1][0].startswith("obs-blank-")
    assert obs_rows[1][7] == "blank"
    assert obs_rows[1][2] == f.id


def test_export_camtrap_dp_keeps_boxes_when_the_file_reads_blank(client, db):
    """Camtrap DP is per box and is the archival export, so it must never
    drop rows it holds. The blank branch used to also fire on the stored
    observation_type, which was near-equivalent while that column was
    derived over every frame. It is not equivalent now: this video reads
    blank because its best frame is empty, yet it still has a passing box
    on frame 50, and that box has to reach the archive."""
    project, _site, deployment = _build_simple_project(db)
    f = make_file(
        db,
        deployment_id=deployment.id,
        captured_at_local=datetime(2024, 6, 15, 9, 0, 0),
        file_type="video",
        file_format="mp4",
        best_frame_number=3,
        observation_type="blank",
    )
    make_detection(
        db,
        file_id=f.id,
        category="animal",
        confidence=0.9,
        frame_number=50,
        label="deer",
    )
    db.commit()

    resp = _run_camtrap_dp_export(client, db, project.id)
    assert resp.status_code == 200
    with zipfile.ZipFile(io.BytesIO(resp.content)) as zf:
        obs_rows = list(csv.reader(io.StringIO(zf.read("observations.csv").decode())))

    media_rows = [r for r in obs_rows[1:] if r[6] == "media"]
    assert media_rows, "the box was dropped from the archival export"
    assert not any(r[0].startswith("obs-blank-") for r in media_rows)


def test_export_camtrap_dp_leaves_out_files_without_date(client, db):
    """CamtrapDP requires a timestamp on media rows and eventStart/eventEnd
    on observation rows, so a file with no capture date cannot be
    represented: an empty string there fails validation in the camtrapdp
    R package and GBIF ingestion, poisoning the whole package. Such
    files are left out; the export dialog warns about them via the
    /files-without-date count."""
    project, _site, deployment = _build_simple_project(db)
    dated = make_file(
        db,
        deployment_id=deployment.id,
        captured_at_local=datetime(2024, 6, 15, 9, 0, 0),
    )
    # make_file coalesces None to a default date, so null it afterwards.
    dateless = make_file(db, deployment_id=deployment.id)
    dateless.captured_at_local = None
    make_detection(db, file_id=dateless.id, category="animal", confidence=0.9)
    db.commit()

    resp = _run_camtrap_dp_export(client, db, project.id)
    assert resp.status_code == 200
    with zipfile.ZipFile(io.BytesIO(resp.content)) as zf:
        media_rows = list(csv.reader(io.StringIO(zf.read("media.csv").decode())))
        obs_rows = list(csv.reader(io.StringIO(zf.read("observations.csv").decode())))

    media_ids = {r[0] for r in media_rows[1:]}
    assert dated.id in media_ids
    assert dateless.id not in media_ids
    # No observation row references the dateless file, and every emitted
    # row has non-empty eventStart / eventEnd.
    assert not any(r[2] == dateless.id for r in obs_rows[1:])
    assert all(r[4] and r[5] for r in obs_rows[1:])


def test_files_without_date_count_endpoint(client, db):
    """The CamtrapDP export dialog warns using this count."""
    project, _site, deployment = _build_simple_project(db)
    make_file(
        db,
        deployment_id=deployment.id,
        captured_at_local=datetime(2024, 6, 15, 9, 0, 0),
    )
    dateless = make_file(db, deployment_id=deployment.id)
    dateless.captured_at_local = None
    db.commit()

    resp = client.get(f"/api/projects/{project.id}/files-without-date")
    assert resp.status_code == 200
    assert resp.json() == {"count": 1}


def test_export_camtrap_dp_emits_media_and_event_rows(client, db):
    """Camtrap-DP dual model: one media-level row per bounding box, plus
    one event-level row per species carrying the effective (human) count
    with no bbox. Replaces the retired box-less observation flow."""
    from app.api.crud.event_observation import (
        calculate_max_n_for_event,
        set_human_count,
    )

    project, _site, deployment = _build_simple_project(db, timezone="UTC")
    ev = make_event_with_files(
        db,
        deployment_id=deployment.id,
        event_start_local=datetime(2024, 6, 15, 9, 0, 0),
    )
    det = make_detection(
        db,
        file_id=ev.files[0].id,
        category="animal",
        confidence=0.9,
        label="deer",
        bbox_x=0.1,
        bbox_y=0.1,
        bbox_width=0.2,
        bbox_height=0.2,
    )
    tax = LabelTaxonomy(
        name="deer", level="species", classification_model_id="",
        project_id=project.id, common_name="Deer",
        scientific_name="Cervidae",
    )
    db.add(tax)
    db.flush()
    det.label_taxonomy_id = tax.id
    # The MaxN rebuild groups by label_taxonomy_id in SQL, so the link has
    # to reach the database first (the session runs autoflush=False, like
    # the app's; every production caller has committed by this point).
    db.flush()
    obs = calculate_max_n_for_event(db, ev.id, project.counting_threshold)
    db.flush()
    # Human bumps the deer count to 3 (more than any single frame showed).
    set_human_count(db, obs[0].id, 3)
    db.commit()

    resp = _run_camtrap_dp_export(client, db, project.id)
    assert resp.status_code == 200
    with zipfile.ZipFile(io.BytesIO(resp.content)) as zf:
        headers, *obs_rows = csv.reader(
            io.StringIO(zf.read("observations.csv").decode())
        )

    level_i = headers.index("observationLevel")
    count_i = headers.index("count")
    bx_i = headers.index("bboxX")
    sci_i = headers.index("scientificName")

    media = [r for r in obs_rows if r[level_i] == "media"]
    event = [r for r in obs_rows if r[level_i] == "event"]
    # One media-level row per box (the deer detection), bbox set, count 1.
    assert len(media) == 1
    assert media[0][bx_i] != ""
    assert media[0][count_i] == "1"
    # One event-level row per species carrying the effective count (3),
    # no bbox, with the resolved scientific name.
    assert len(event) == 1
    assert event[0][count_i] == "3"
    assert event[0][bx_i] == ""
    assert event[0][sci_i] == "Cervidae"


# ---------------------------------------------------------------------------
# Unit tests for the pure serializers
# ---------------------------------------------------------------------------


def test_slugify_edges():
    assert export_formats.slugify("My Project!") == "my-project"
    assert export_formats.slugify("under_score name") == "under-score-name"
    assert export_formats.slugify("   ") == "project"


def test_make_gpkg_point_blob_layout():
    blob = export_formats.make_gpkg_point_blob(5.1, 52.1)
    # Header: 'GP' + version(0) + flags(1) + srid(4326, LE int32) = 8 bytes
    assert blob[:2] == b"GP"
    assert blob[2] == 0
    assert blob[3] == 1
    assert int.from_bytes(blob[4:8], "little", signed=True) == 4326
    # WKB Point: byte-order(1) + type(1) + X + Y = 1+4+8+8 = 21 bytes
    assert len(blob) == 8 + 21
    assert blob[8] == 1
    assert int.from_bytes(blob[9:13], "little") == 1


def test_serialize_csv_roundtrip():
    payload = export_formats.serialize_csv(
        ["a", "b"], [[1, "x"], [2, "y,z"]]
    )
    rows = list(csv.reader(io.StringIO(payload.decode("utf-8"))))
    assert rows == [["a", "b"], ["1", "x"], ["2", "y,z"]]


def test_scoped_rows_defer_exif_blob(db):
    """get_scoped_detection_rows must not pull File.exif_data into the sort.

    The export never reads the per-file EXIF JSON blob, so it is deferred.
    Loading it dragged ~70k blobs through SQLite's ORDER BY sorter and
    blew up the temp file with SQLITE_FULL on large projects. Asserting
    the column is unloaded pins the fix in place.
    """
    from sqlalchemy import inspect

    from app.api.crud.export import get_scoped_detection_rows

    project = make_project(db)
    site = make_site(db, project_id=project.id)
    deployment = make_deployment(db, project_id=project.id, site_id=site.id)
    file_obj = make_file(
        db,
        deployment_id=deployment.id,
        exif_data={"Make": "RECONYX", "huge": "x" * 5000},
    )
    make_detection(db, file_id=file_obj.id, category="animal", confidence=0.9)

    # Expire so the query reloads from the DB; otherwise the freshly built
    # instance sits in the identity map with exif_data already populated and
    # defer() has nothing to unload.
    db.expire_all()

    rows = get_scoped_detection_rows(db, project)

    assert rows, "expected at least one scoped row"
    returned_file = rows[0][0]
    # Deferred: exif_data is not loaded until explicitly touched.
    assert "exif_data" in inspect(returned_file).unloaded
    # But it is still reachable when something does ask for it.
    assert returned_file.exif_data["Make"] == "RECONYX"


def test_build_files_rows_accepts_prefetched_rows(db):
    """Passing the default-query scoped rows must yield exactly the rows
    the self-fetching call produces. Pins the reuse path that
    build_spreadsheet_sheets takes to avoid running the same query twice."""
    from app.api.crud.export import (
        build_files_rows,
        get_scoped_detection_rows,
    )

    project = make_project(db)
    site = make_site(db, project_id=project.id)
    deployment = make_deployment(db, project_id=project.id, site_id=site.id)
    with_box = make_file(db, deployment_id=deployment.id)
    make_detection(db, file_id=with_box.id, category="animal", confidence=0.9)
    make_file(db, deployment_id=deployment.id)  # empty file, blank row

    headers_direct, rows_direct = build_files_rows(db, project)
    scoped = get_scoped_detection_rows(db, project)
    headers_reused, rows_reused = build_files_rows(
        db, project, scoped_rows=scoped
    )

    assert headers_reused == headers_direct
    assert rows_reused == rows_direct
    assert len(rows_reused) == 2


def test_serialize_xlsx_multi_write_only_roundtrip():
    """write_only mode still produces a readable multi-sheet workbook.

    Pins the memory-bounded serializer: rows stream into the sheet rather
    than building an in-memory cell graph, but the output must load back
    with the same sheet names and cell values.
    """
    from openpyxl import load_workbook

    payload = export_formats.serialize_xlsx_multi(
        [
            ("Alpha", ["a", "b"], [[1, "x"], [2, "y"]]),
            ("Beta", ["c"], [["only"]]),
        ]
    )
    wb = load_workbook(io.BytesIO(payload))
    assert wb.sheetnames == ["Alpha", "Beta"]
    assert list(wb["Alpha"].iter_rows(values_only=True)) == [
        ("a", "b"),
        (1, "x"),
        (2, "y"),
    ]
    assert list(wb["Beta"].iter_rows(values_only=True)) == [
        ("c",),
        ("only",),
    ]


def test_write_xlsx_multi_writes_same_workbook_to_disk(tmp_path):
    """The path-writing variant (used by the folder-run save so the
    zipped workbook never exists as one in-memory blob) must produce
    the same content as the bytes serializer."""
    from openpyxl import load_workbook

    sheets = [
        ("Alpha", ["a", "b"], [[1, "x"], [2, "y"]]),
        ("Beta", ["c"], [["only"]]),
    ]
    target = tmp_path / "out.xlsx"
    export_formats.write_xlsx_multi(sheets, target)

    wb = load_workbook(target)
    assert wb.sheetnames == ["Alpha", "Beta"]
    assert list(wb["Alpha"].iter_rows(values_only=True)) == [
        ("a", "b"),
        (1, "x"),
        (2, "y"),
    ]


def test_events_by_file_survives_bound_parameter_limit(db):
    """Regression: large folder runs crashed the save step with
    "too many SQL variables" because `_events_by_file` built one
    `IN (?, ?, ...)` over every scoped file id (Simon's 45k-file run).

    Reproduced at small scale by lowering SQLite's bound-parameter limit
    on the shared test connection to the old-build value (999). With more
    than 999 files in one event scope, the un-chunked query raises; the
    chunked helper (900 per batch) must return the full mapping instead.
    """
    import sqlite3

    from app.api.crud.export import _events_by_file
    from app.db.sql_params import SQL_VAR_CHUNK

    project = make_project(db)
    site = make_site(db, project_id=project.id)
    deployment = make_deployment(db, site_id=site.id, project_id=project.id)

    # One event holding more files than the (lowered) parameter limit, so the
    # id list spans more than one chunk.
    n_files = SQL_VAR_CHUNK + 200
    event = make_event_with_files(
        db,
        deployment_id=deployment.id,
        event_start_local=datetime(2024, 1, 1, 12, 0, 0),
        files_verified=[False] * n_files,
    )
    file_ids = [
        fid
        for (fid,) in db.query(File.id)
        .filter(File.deployment_id == deployment.id)
        .all()
    ]
    assert len(file_ids) == n_files

    raw = db.connection().connection
    raw = getattr(raw, "dbapi_connection", raw)
    previous = raw.setlimit(sqlite3.SQLITE_LIMIT_VARIABLE_NUMBER, 999)
    try:
        result = _events_by_file(db, file_ids)
    finally:
        raw.setlimit(sqlite3.SQLITE_LIMIT_VARIABLE_NUMBER, previous)

    assert len(result) == n_files
    assert all(ev.id == event.id for ev in result.values())


# ---------------------------------------------------------------------------
# Camtrap DP is an external standard: observationType has a fixed vocabulary
# ---------------------------------------------------------------------------


def test_camtrap_observation_type_translates_raw_categories():
    """`observation_type` and `Detection.category` carry the detector's own
    words everywhere in the app, so a marine model's `shark` reaches the
    folder tree and the generic CSV intact. Camtrap DP cannot take it: the
    standard has no marine categories and expects all wildlife under
    `animal` with the species in `scientificName`. This function is the
    only place that translation happens."""
    from app.api.crud.export import (
        CAMTRAP_OBSERVATION_TYPES,
        _obs_type_from_category,
    )

    # MegaDetector's three, including the one that is renamed.
    assert _obs_type_from_category("animal") == "animal"
    assert _obs_type_from_category("person") == "human"
    assert _obs_type_from_category("vehicle") == "vehicle"
    assert _obs_type_from_category("blank") == "blank"

    # Anything else is wildlife. No list of marine categories to keep in
    # step with the detectors.
    for category in ("shark", "fish", "turtle", "something_new"):
        assert _obs_type_from_category(category) == "animal"

    # Whatever comes in, what comes out is always in the standard's
    # vocabulary. This is the assertion that protects the export.
    for category in (
        "animal", "person", "vehicle", "blank", "shark", "fish", "",
    ):
        assert _obs_type_from_category(category) in CAMTRAP_OBSERVATION_TYPES


def test_spatial_detection_count_ignores_off_best_frame_boxes(client, db):
    """A map bubble must count what the Labels grid holds. Only one frame
    per video is written to disk, so boxes on the other frames have no
    picture to open; counting them made the deployments layer report 220
    where detections.csv listed 32."""
    project, _site, deployment = _build_simple_project(db)
    video = make_file(
        db,
        deployment_id=deployment.id,
        file_type="video",
        file_format="mp4",
        best_frame_number=3,
    )
    for frame in (3, 7, 11):
        make_detection(
            db, file_id=video.id, category="animal", confidence=0.9,
            label="fox", frame_number=frame,
        )
    db.commit()

    resp = client.get(f"/api/projects/{project.id}/export/spatial?format=geojson")
    assert resp.status_code == 200
    payload = json.loads(resp.content)
    dep_feature = next(
        feat for feat in payload["features"]
        if feat["properties"]["layer"] == "deployments"
    )
    assert dep_feature["properties"]["detection_count"] == 1


def test_export_camtrap_dp_variant_rows(client, db):
    """A variant class exports the plain binomial as scientificName; the
    variant itself fills lifeStage or sex when it fits the standard's
    vocabulary and rides in observationComments otherwise. The taxonomic
    scope deduplicates variants into one species entry."""
    project, _site, deployment = _build_simple_project(
        db, timezone="Europe/Amsterdam"
    )

    def _variant_row(variant: str) -> LabelTaxonomy:
        row = LabelTaxonomy(
            id=str(uuid.uuid4()),
            classification_model_id="TEST-MODEL",
            name=f"red fox {variant}",
            level="variant",
            taxon_class="mammalia",
            taxon_order="carnivora",
            taxon_family="canidae",
            taxon_genus="vulpes",
            taxon_species="vulpes",
            taxon_variant=variant,
            common_name=f"Red fox {variant}",
            scientific_name=f"V. vulpes ({variant})",
        )
        db.add(row)
        db.flush()
        return row

    for variant in ("adult", "male", "melanistic"):
        taxonomy = _variant_row(variant)
        f = make_file(
            db,
            deployment_id=deployment.id,
            captured_at_local=datetime(2024, 6, 15, 9, 0, 0),
        )
        make_detection(
            db,
            file_id=f.id,
            category="animal",
            confidence=0.9,
            label=taxonomy.name,
            scientific_name=taxonomy.scientific_name,
            label_confidence=0.88,
            label_taxonomy_id=taxonomy.id,
        )
    db.commit()

    resp = _run_camtrap_dp_export(client, db, project.id)
    assert resp.status_code == 200
    with zipfile.ZipFile(io.BytesIO(resp.content)) as zf:
        dp = json.loads(zf.read("datapackage.json"))
        obs_rows = list(
            csv.reader(io.StringIO(zf.read("observations.csv").decode()))
        )

    header = obs_rows[0]
    i_sci = header.index("scientificName")
    i_life = header.index("lifeStage")
    i_sex = header.index("sex")
    i_comments = header.index("observationComments")

    data = obs_rows[1:]
    assert len(data) == 3
    # Every row carries the real binomial, never the qualified leaf name.
    assert {r[i_sci] for r in data} == {"V. vulpes"}

    by_life = {r[i_life] for r in data}
    by_sex = {r[i_sex] for r in data}
    assert "adult" in by_life
    assert "male" in by_sex
    # The non-enum variant survives in the comments instead.
    assert any("variant: melanistic" in r[i_comments] for r in data)
    # An enum-mapped variant does not leak into the comments.
    assert not any("variant: adult" in r[i_comments] for r in data)

    # One species entry in the taxonomic scope, at species rank.
    fox_entries = [
        e for e in dp["taxonomic"] if e["scientificName"] == "V. vulpes"
    ]
    assert len(fox_entries) == 1
    assert fox_entries[0]["taxonRank"] == "species"


# ---------------------------------------------------------------------------
# Paired cameras in the deployment exports
# ---------------------------------------------------------------------------


def test_export_deployments_rows_carry_paired_cameras(db):
    from app.api.crud.export import build_deployments_rows

    project, _site, deployment = _build_simple_project(db)
    deployment.paired_cameras = True
    db.commit()

    headers, rows = build_deployments_rows(db, project)
    assert "paired_cameras" in headers
    assert rows[0][headers.index("paired_cameras")] == "true"


def test_export_camtrap_dp_tags_paired_cameras(client, db):
    """Camtrap DP defines a deployment as one camera, so a paired
    deployment says so in deploymentTags, next to the user's own tags."""
    project, _site, deployment = _build_simple_project(db, timezone="Europe/Amsterdam")
    deployment.paired_cameras = True
    deployment.tags = {"season": "wet"}
    make_file(db, deployment_id=deployment.id, captured_at_local=datetime(2024, 6, 15, 9, 0, 0))
    db.commit()

    resp = _run_camtrap_dp_export(client, db, project.id)
    assert resp.status_code == 200
    with zipfile.ZipFile(io.BytesIO(resp.content)) as zf:
        deps_rows = list(csv.reader(io.StringIO(zf.read("deployments.csv").decode())))
    tags_col = deps_rows[0].index("deploymentTags")
    assert deps_rows[1][tags_col] == "season:wet | paired_cameras:true"
